import json
import csv
import openpyxl

from mytools.tool_class import CodeNamePhone
from openpyxl import Workbook

def load_models_csv()->dict:
    new_models = CodeNamePhone("models.csv").get_names_code_csv
    new_models = {value: [key] for key, value in new_models.items()}
    print(new_models)
    old_models = CodeNamePhone("old_models.csv").get_names_code_csv
    old_models = {value: [key] for key, value in old_models.items()}
    print(old_models)

    all_models = old_models.copy()

    for key, value in new_models.items():
        all_models[key] = all_models.get(key,[]) + value if value != all_models.get(key,[]) else all_models[key]

    print(all_models)

    return all_models

def load_models_xlsx()->dict:
    workbook = openpyxl.load_workbook("models.xlsx")
    sheet = workbook.active

    data = {}
    for row in sheet.iter_rows(min_row=2):
        key = row[0].value
        value = row[1].value.split(",")
        if key:
            data[key] = value

    return data

def save_to_excel(models_dict: dict):
    # Создаем новую рабочую книгу
    wb = Workbook()

    # Выбираем активный лист
    sheet = wb.active

    # Задаем заголовки столбцов
    sheet['A1'] = 'name'
    sheet['B1'] = 'code'

    # Заполняем столбики ключами и значениями
    row = 2
    for key, value in sorted(models_dict.items()):
        val_str = ",".join(value)
        sheet.cell(row=row, column=1, value=key)
        sheet.cell(row=row, column=2, value=val_str)
        row += 1

    # Сохраняем рабочую книгу в файл
    wb.save('models11.xlsx')
    print('Словарь успешно сохранен в xlsx файл!')

def save_to_csv(models_dict: dict):
    # Открытие файла для записи
    with open("all_models.csv", "w", encoding="utf-16", newline="") as csv_file:
        # Сохранение словаря в CSV файл
        writer = csv.writer(csv_file, delimiter=":")
        writer.writerows(sorted(models_dict.items()))

    print('Словарь успешно сохранен в csv файл!')

def save_to_json(models_dict: dict):
    # Открытие файла для записи
    with open("all_models.json", "w", encoding="utf-8", newline="") as json_file:
        # Сохранение словаря в JSON файл
        json.dump(models_dict, json_file, indent=4, separators=(",", ":"), sort_keys=models_dict.keys())

    print('Словарь успешно сохранен в json файл!')


if __name__ == '__main__':
    # models_dict = load_models_csv()
    models_dict = load_models_xlsx()
    # save_to_json(models_dict)
    # save_to_csv(models_dict)
    save_to_excel(models_dict)


    # pip freeze > requirements.txt
    # pip install -r requirements.txt